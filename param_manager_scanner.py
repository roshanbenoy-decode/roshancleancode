"""
ParamManager Scanner - Scan Azure Blob Storage for ParamManager files in Datafeed folders.

This script searches through the '30-projects' container for folders named 'Datafeed',
analyzes Excel files (ParamManager.xlsx) and Parquet files within them, and exports
metadata (table names and column names) to a CSV report.
"""

import os
import sys
import tempfile
import time
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from azure.core.exceptions import AzureError
from azure_file_manager import AzureFileManager


class ParamManagerScanner(AzureFileManager):
    """Scans Azure Blob Storage for ParamManager files in Datafeed folders.

    Inherits from AzureFileManager to reuse authentication and blob operations.
    """

    EXCEL_SEARCH_TERM = "parametermanager"  # Search for files containing this term (lowercase for comparison)

    def __init__(self):
        """Initialize the ParamManager Scanner by calling parent AzureFileManager."""
        # Call parent class initialization (handles all authentication)
        super().__init__()

    def scan_for_datafeed_folders(self):
        """Scan all blobs and identify Datafeed folders.

        Returns:
            list: List of unique Datafeed folder paths
        """
        try:
            print("Scanning container for 'Datafeed' folders...")
            print("-" * 80)

            container_client = self.blob_service_client.get_container_client(self.CONTAINER_NAME)
            blobs = container_client.list_blobs()

            datafeed_folders = set()

            for blob in blobs:
                # Check if path contains 'Datafeed' as a folder
                path_parts = blob.name.split('/')

                # Look for 'Datafeed' folder in the path
                for i, part in enumerate(path_parts):
                    if part.lower() == 'datafeed':
                        # Construct the path up to and including 'Datafeed'
                        datafeed_path = '/'.join(path_parts[:i+1]) + '/'
                        datafeed_folders.add(datafeed_path)
                        break

            sorted_folders = sorted(list(datafeed_folders))
            print(f"✓ Found {len(sorted_folders)} Datafeed folder(s)\n")

            if sorted_folders:
                print("Datafeed folders found:")
                for folder in sorted_folders:
                    print(f"  - {folder}")
                print()

            return sorted_folders

        except AzureError as e:
            print(f"Error scanning blobs: {e}")
            return []

    def get_files_in_datafeed(self, datafeed_path):
        """Get ParamManager Excel file in a specific Datafeed folder.

        Args:
            datafeed_path: Path to the Datafeed folder

        Returns:
            str: Path to Excel file, or None if not found
        """
        try:
            container_client = self.blob_service_client.get_container_client(self.CONTAINER_NAME)
            blobs = container_client.list_blobs(name_starts_with=datafeed_path)

            for blob in blobs:
                # Skip folder markers
                if blob.size == 0 or blob.name.endswith('/'):
                    continue

                # Get filename
                filename = blob.name.split('/')[-1]

                # Check for Excel files containing the search term (e.g., ParameterManager)
                if self.EXCEL_SEARCH_TERM in filename.lower() and filename.lower().endswith(('.xlsx', '.xlsm', '.xls')):
                    return blob.name

            return None

        except AzureError as e:
            print(f"Error listing files in {datafeed_path}: {e}")
            return None

    def download_blob_to_temp(self, blob_name):
        """Download a blob to a temporary file.

        Args:
            blob_name: Full path to the blob

        Returns:
            str: Path to temporary file, or None if failed
        """
        try:
            # Create temporary file
            suffix = Path(blob_name).suffix
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
            temp_path = temp_file.name
            temp_file.close()

            # Download blob
            blob_client = self.blob_service_client.get_blob_client(
                container=self.CONTAINER_NAME,
                blob=blob_name
            )

            with open(temp_path, "wb") as file:
                download_stream = blob_client.download_blob()
                file.write(download_stream.readall())

            return temp_path

        except Exception as e:
            print(f"  ✗ Error downloading {blob_name}: {e}")
            return None

    def analyze_excel_file_with_tables(self, excel_path, datafeed_path):
        """Analyze Excel file and extract named tables with columns (Option B format).

        Args:
            excel_path: Path to the downloaded Excel file
            datafeed_path: Path to the Datafeed folder (for reporting)

        Returns:
            list: List of dictionaries with metadata (one row per column)
        """
        results = []

        try:
            # Load workbook with openpyxl
            wb = load_workbook(excel_path, data_only=True)

            print(f"  ✓ Found {len(wb.sheetnames)} sheet(s) in Excel file")

            total_tables = 0
            total_columns = 0

            # Iterate through all sheets
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Check if sheet has named tables
                if hasattr(ws, 'tables') and ws.tables:
                    # Sheet has named tables
                    for table_name, table_ref in ws.tables.items():
                        total_tables += 1

                        try:
                            # table_ref is already a string (e.g., "A1:D10")
                            # Parse the table range to get cell boundaries
                            min_col, min_row, max_col, max_row = range_boundaries(table_ref)

                            # Extract column headers (first row of the table)
                            column_headers = []
                            for col in range(min_col, max_col + 1):
                                cell_value = ws.cell(row=min_row, column=col).value
                                if cell_value:
                                    column_headers.append(cell_value)

                            # Create one row per column (Option B)
                            for column_name in column_headers:
                                total_columns += 1
                                results.append({
                                    'Path': datafeed_path.rstrip('/'),
                                    'Source_Type': 'Excel',
                                    'Sheet_Name': sheet_name,
                                    'Table_Name': table_name,
                                    'Column_Name': str(column_name)
                                })

                        except Exception as e:
                            print(f"  ✗ Error reading table '{table_name}' in sheet '{sheet_name}': {e}")

                else:
                    # No named tables found in this sheet
                    print(f"  ⚠ Sheet '{sheet_name}' has no named Excel tables")

            wb.close()

            if total_tables > 0:
                print(f"  ✓ Found {total_tables} named table(s) with {total_columns} total columns")
            else:
                print(f"  ⚠ No named Excel tables found in file")

        except Exception as e:
            print(f"  ✗ Error analyzing Excel file: {e}")

        return results


    def generate_report(self):
        """Generate comprehensive report of all ParamManager files in Datafeed folders.

        Returns:
            pd.DataFrame: DataFrame with all metadata
        """
        # Find all Datafeed folders
        datafeed_folders = self.scan_for_datafeed_folders()

        if not datafeed_folders:
            print("No Datafeed folders found in the container.")
            return pd.DataFrame()

        # Collect all metadata
        all_results = []

        print("Analyzing Datafeed folders for ParameterManager files...")
        print("=" * 80)

        for idx, datafeed_path in enumerate(datafeed_folders, 1):
            print(f"\n[{idx}/{len(datafeed_folders)}] Processing: {datafeed_path}")
            print("-" * 80)

            # Get ParamManager Excel file in this Datafeed folder
            excel_file = self.get_files_in_datafeed(datafeed_path)

            # Analyze Excel file if exists
            if excel_file:
                excel_filename = excel_file.split('/')[-1]
                print(f"  Found ParameterManager Excel file: {excel_filename}")
                temp_excel_path = self.download_blob_to_temp(excel_file)

                if temp_excel_path:
                    try:
                        excel_results = self.analyze_excel_file_with_tables(temp_excel_path, datafeed_path)
                        all_results.extend(excel_results)
                    finally:
                        # Clean up temp file - wait a bit for Windows to release file handle
                        if os.path.exists(temp_excel_path):
                            time.sleep(0.1)
                            try:
                                os.unlink(temp_excel_path)
                            except PermissionError:
                                # If still locked, try again after a longer wait
                                time.sleep(0.5)
                                try:
                                    os.unlink(temp_excel_path)
                                except PermissionError:
                                    print(f"  Warning: Could not delete temp file {temp_excel_path}")
            else:
                print(f"  ✗ No Excel file containing '{self.EXCEL_SEARCH_TERM}' found")

        print("\n" + "=" * 80)
        print(f"✓ Scan complete! Processed {len(datafeed_folders)} Datafeed folder(s)")
        print(f"✓ Extracted {len(all_results)} column entries (one row per column)")
        print("=" * 80)

        # Create DataFrame
        if all_results:
            df = pd.DataFrame(all_results)
            return df
        else:
            return pd.DataFrame()

    def export_to_csv(self, df, output_filename=None):
        """Export DataFrame to CSV.

        Args:
            df: DataFrame to export
            output_filename: Optional custom filename
        """
        if df.empty:
            print("\nNo data to export.")
            return

        # Ask user for output location (following project guidelines)
        print("\n" + "=" * 80)
        print("OUTPUT FILE LOCATION")
        print("=" * 80)
        output_dir = input("Enter output directory (press Enter for default 'downloads'): ").strip()

        # Use downloads as default if no input provided
        if not output_dir:
            output_dir = 'downloads'

        # Create directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)

        # Generate filename with timestamp if not provided
        if not output_filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"parammanager_report_{timestamp}.csv"

        # Build full output path
        output_path = os.path.join(output_dir, output_filename)

        try:
            df.to_csv(output_path, index=False)
            print(f"\n✓ Report exported to: {output_path}")
            print(f"  Total rows: {len(df)}")
            print(f"  Columns: {', '.join(df.columns.tolist())}")
        except Exception as e:
            print(f"\n✗ Error exporting CSV: {e}")


def main():
    """Main function."""
    print("=" * 80)
    print("ParameterManager Scanner - Azure Blob Storage")
    print("Container: 30-projects")
    print("Searching for: ParameterManager files in Datafeed folders")
    print("=" * 80)
    print()

    # Initialize scanner
    scanner = ParamManagerScanner()

    # Generate report
    df = scanner.generate_report()

    # Display summary
    if not df.empty:
        print("\n" + "=" * 80)
        print("REPORT PREVIEW")
        print("=" * 80)
        print(df.to_string(index=False, max_rows=10))

        if len(df) > 10:
            print(f"\n... ({len(df) - 10} more rows)")

        print("\n" + "=" * 80)
        print("SUMMARY STATISTICS")
        print("=" * 80)
        print(f"Total column entries: {len(df)}")
        print(f"Unique Datafeed paths: {df['Path'].nunique()}")
        print(f"ParameterManager Excel named tables: {df['Table_Name'].nunique()}")
        print(f"Total columns extracted: {len(df)}")
        print("=" * 80)

        # Export to CSV
        scanner.export_to_csv(df)
    else:
        print("\nNo ParameterManager data found to report.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user. Goodbye!")
        sys.exit(0)
